#!/usr/bin/env python3
"""
seed_from_xlsx.py
─────────────────
Seeds the Sales Operating System SQLite database from Sales_Operating_System.xlsx.

Populates:
  - product_catalog     (from Master_Catalog sheet)
  - upsell_rules        (from Upsell_Matrix sheet)
  - cross_sell_rules    (from Cross_Sell_Matrix sheet)
  - bundles             (from Bundles sheet)
  - bundle_items        (from Bundles sheet, normalized)
  - need_states         (from Client_Needs_Mapping sheet)
  - need_state_products (from Client_Needs_Mapping, product references resolved)

Usage:
    python scripts/seed_from_xlsx.py [--xlsx PATH] [--db PATH] [--clear]

Options:
    --xlsx   Path to Sales_Operating_System.xlsx
             (default: ~/Desktop/Sales_Operating_System.xlsx)
    --db     Path to SQLite database file
             (default: sales_os.db in project root)
    --clear  Drop and re-seed all seeded tables before inserting

Idempotent by default — skips rows where the ID already exists.
Use --clear to fully re-seed from scratch.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# Utilities
# ─────────────────────────────────────────────────────────────────────────────

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _uid() -> str:
    return str(uuid.uuid4())


def _cell(val: Any) -> Optional[str]:
    if val is None:
        return None
    v = str(val).strip()
    return v if v else None


def _int_cell(val: Any, default: int = 1) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _float_cell(val: Any, default: float = 0.0) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _sheet_rows(wb: openpyxl.Workbook, sheet_name: str) -> Tuple[List[str], List[Dict]]:
    """Return (headers, list_of_dicts) for a sheet, skipping empty rows."""
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    headers: List[str] = []
    rows: List[Dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not any(c for c in row):
            continue
        if not headers:
            headers = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(row)]
            continue
        row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
        rows.append(row_dict)
    return headers, rows


# ─────────────────────────────────────────────────────────────────────────────
# Slugify product name → stable ID for cross-referencing
# ─────────────────────────────────────────────────────────────────────────────

def _name_to_slug(name: str) -> str:
    """Convert product name to a stable lowercase slug for FK resolution."""
    return name.lower().strip().replace(" ", "-").replace("/", "-").replace("&", "and")


# ─────────────────────────────────────────────────────────────────────────────
# Seeders
# ─────────────────────────────────────────────────────────────────────────────

def seed_catalog(
    conn: sqlite3.Connection,
    wb: openpyxl.Workbook,
    name_to_id: Dict[str, str],
) -> None:
    """Seed product_catalog from Master_Catalog sheet."""
    _, rows = _sheet_rows(wb, "Master_Catalog")
    inserted = 0
    skipped = 0

    for row in rows:
        xlsx_id = _cell(row.get("ID"))
        name = _cell(row.get("Name"))
        if not name:
            continue

        product_id = _uid()
        name_to_id[_name_to_slug(name)] = product_id
        # Also map by xlsx ID for upsell/cross-sell resolution
        if xlsx_id:
            name_to_id[xlsx_id.lower()] = product_id

        # Check if already exists by name
        existing = conn.execute(
            "SELECT id FROM product_catalog WHERE name = ?", (name,)
        ).fetchone()
        if existing:
            name_to_id[_name_to_slug(name)] = existing[0]
            if xlsx_id:
                name_to_id[xlsx_id.lower()] = existing[0]
            skipped += 1
            continue

        conn.execute(
            """
            INSERT INTO product_catalog
              (id, name, type, category, subcategory, description, primary_goal,
               core_value, interaction_value, marketing_influence, score_multiplier,
               funnel_stage, primary_channel, automation_potential, source_reference,
               is_active, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?)
            """,
            (
                product_id,
                name,
                _cell(row.get("Type")),
                _cell(row.get("Category")),
                _cell(row.get("Subcategory")),
                _cell(row.get("Description")),
                _cell(row.get("Primary_Goal")),
                _cell(row.get("Core_Value")),
                _int_cell(row.get("Interaction_Value (1-5)"), 3),
                _int_cell(row.get("Marketing_Influence (1-5)"), 3),
                _float_cell(row.get("Score_Multiplier"), 1.0),
                _cell(row.get("Funnel_Stage")),
                _cell(row.get("Primary_Channel")),
                _cell(row.get("Automation_Potential")),
                _cell(row.get("Source_Reference")),
                _now(),
                _now(),
            ),
        )
        inserted += 1

    conn.commit()
    print(f"  product_catalog: {inserted} inserted, {skipped} skipped")


def seed_upsell_rules(
    conn: sqlite3.Connection,
    wb: openpyxl.Workbook,
    name_to_id: Dict[str, str],
) -> None:
    """Seed upsell_rules from Upsell_Matrix sheet."""
    # Ensure table exists
    conn.execute("""
        CREATE TABLE IF NOT EXISTS upsell_rules (
            id                    TEXT PRIMARY KEY,
            primary_product_id    TEXT,
            client_need_state_id  TEXT,
            trigger_event         TEXT,
            recommended_product_id TEXT,
            upsell_type           TEXT,
            expected_impact       TEXT,
            dependency_product_id TEXT,
            priority_order        INTEGER DEFAULT 5,
            created_at            TEXT
        )
    """)

    _, rows = _sheet_rows(wb, "Upsell_Matrix")
    inserted = 0

    for row in rows:
        primary_name = _cell(row.get("Primary_Product"))
        rec_name = _cell(row.get("Recommended_Upsell"))
        dep_name = _cell(row.get("Dependency"))

        if not primary_name or not rec_name:
            continue

        primary_id = name_to_id.get(_name_to_slug(primary_name))
        rec_id = name_to_id.get(_name_to_slug(rec_name))
        dep_id = name_to_id.get(_name_to_slug(dep_name)) if dep_name else None

        conn.execute(
            """
            INSERT INTO upsell_rules
              (id, primary_product_id, trigger_event, recommended_product_id,
               upsell_type, expected_impact, dependency_product_id)
            VALUES (?,?,?,?,?,?,?)
            """,
            (
                _uid(),
                primary_id,
                _cell(row.get("Trigger_Event")),
                rec_id,
                _cell(row.get("Upsell_Type")),
                _cell(row.get("Expected_Impact")),
                dep_id,
            ),
        )
        inserted += 1

    conn.commit()
    print(f"  upsell_rules: {inserted} inserted")


def seed_cross_sell_rules(
    conn: sqlite3.Connection,
    wb: openpyxl.Workbook,
    name_to_id: Dict[str, str],
) -> None:
    """Seed cross_sell_rules from Cross_Sell_Matrix sheet."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cross_sell_rules (
            id                TEXT PRIMARY KEY,
            product_id        TEXT,
            paired_product_id TEXT,
            reason            TEXT,
            bundle_strength   INTEGER DEFAULT 3,
            created_at        TEXT
        )
    """)

    _, rows = _sheet_rows(wb, "Cross_Sell_Matrix")
    inserted = 0

    for row in rows:
        prod_name = _cell(row.get("Product"))
        paired_name = _cell(row.get("Pairs_With"))
        if not prod_name or not paired_name:
            continue

        prod_id = name_to_id.get(_name_to_slug(prod_name))
        paired_id = name_to_id.get(_name_to_slug(paired_name))

        conn.execute(
            """
            INSERT INTO cross_sell_rules
              (id, product_id, paired_product_id, reason, bundle_strength)
            VALUES (?,?,?,?,?)
            """,
            (
                _uid(),
                prod_id,
                paired_id,
                _cell(row.get("Reason")),
                _int_cell(row.get("Bundle_Strength (1-5)"), 3),
            ),
        )
        inserted += 1

    conn.commit()
    print(f"  cross_sell_rules: {inserted} inserted")


def seed_bundles_and_items(
    conn: sqlite3.Connection,
    wb: openpyxl.Workbook,
    name_to_id: Dict[str, str],
) -> None:
    """Seed bundles and bundle_items from Bundles sheet."""
    _, rows = _sheet_rows(wb, "Bundles")
    bundle_name_to_id: Dict[str, str] = {}
    bundle_sequence: Dict[str, int] = {}
    bundles_inserted = 0
    items_inserted = 0

    for row in rows:
        bundle_name = _cell(row.get("Bundle_Name"))
        component_name = _cell(row.get("Component"))
        if not bundle_name or not component_name:
            continue

        # Create bundle if new
        if bundle_name not in bundle_name_to_id:
            existing = conn.execute(
                "SELECT id FROM bundles WHERE name = ?", (bundle_name,)
            ).fetchone()
            if existing:
                bundle_name_to_id[bundle_name] = existing[0]
            else:
                bundle_id = _uid()
                target_need = _cell(row.get("Target_Need"))
                conn.execute(
                    """
                    INSERT INTO bundles (id, name, value_proposition, created_at, updated_at)
                    VALUES (?,?,?,?,?)
                    """,
                    (bundle_id, bundle_name, target_need, _now(), _now()),
                )
                bundle_name_to_id[bundle_name] = bundle_id
                bundle_sequence[bundle_name] = 0
                bundles_inserted += 1

        bundle_id = bundle_name_to_id[bundle_name]
        product_id = name_to_id.get(_name_to_slug(component_name))
        if not product_id:
            continue

        seq = bundle_sequence.get(bundle_name, 0)
        bundle_sequence[bundle_name] = seq + 1

        # Skip duplicate bundle_items
        existing_item = conn.execute(
            "SELECT id FROM bundle_items WHERE bundle_id = ? AND product_id = ?",
            (bundle_id, product_id),
        ).fetchone()
        if existing_item:
            continue

        conn.execute(
            """
            INSERT INTO bundle_items (id, bundle_id, product_id, sequence_order, required)
            VALUES (?,?,?,?,1)
            """,
            (_uid(), bundle_id, product_id, seq),
        )
        items_inserted += 1

    conn.commit()
    print(f"  bundles: {bundles_inserted} inserted")
    print(f"  bundle_items: {items_inserted} inserted")


def seed_need_states(
    conn: sqlite3.Connection,
    wb: openpyxl.Workbook,
    name_to_id: Dict[str, str],
) -> None:
    """Seed need_states and need_state_products from Client_Needs_Mapping sheet."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS need_states (
            id              TEXT PRIMARY KEY,
            problem_name    TEXT,
            detected_signal TEXT,
            severity        TEXT,
            description     TEXT,
            priority_order  INTEGER DEFAULT 2,
            created_at      TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS need_state_products (
            id             TEXT PRIMARY KEY,
            need_state_id  TEXT NOT NULL,
            product_id     TEXT NOT NULL,
            priority_order INTEGER DEFAULT 1
        )
    """)

    _, rows = _sheet_rows(wb, "Client_Needs_Mapping")
    ns_inserted = 0
    nsp_inserted = 0

    for row in rows:
        problem = _cell(row.get("Client_Problem"))
        signal = _cell(row.get("Detected_Signal"))
        products_raw = _cell(row.get("Recommended_Products"))
        priority = _int_cell(row.get("Priority_Order"), 2)

        if not problem:
            continue

        existing = conn.execute(
            "SELECT id FROM need_states WHERE problem_name = ?", (problem,)
        ).fetchone()
        if existing:
            ns_id = existing[0]
        else:
            ns_id = _uid()
            conn.execute(
                """
                INSERT INTO need_states (id, problem_name, detected_signal)
                VALUES (?,?,?)
                """,
                (ns_id, problem, signal),
            )
            ns_inserted += 1

        if products_raw:
            for prod_name in [p.strip() for p in products_raw.split(";")]:
                if not prod_name:
                    continue
                prod_id = name_to_id.get(_name_to_slug(prod_name))
                if not prod_id:
                    continue
                existing_link = conn.execute(
                    "SELECT id FROM need_state_products WHERE need_state_id = ? AND product_id = ?",
                    (ns_id, prod_id),
                ).fetchone()
                if existing_link:
                    continue
                conn.execute(
                    """
                    INSERT INTO need_state_products (id, need_state_id, product_id, priority_order)
                    VALUES (?,?,?,?)
                    """,
                    (_uid(), ns_id, prod_id, priority),
                )
                nsp_inserted += 1

    conn.commit()
    print(f"  need_states: {ns_inserted} inserted")
    print(f"  need_state_products: {nsp_inserted} inserted")


# ─────────────────────────────────────────────────────────────────────────────
# Clear helpers
# ─────────────────────────────────────────────────────────────────────────────

SEEDED_TABLES = [
    "need_state_products",
    "need_states",
    "bundle_items",
    "bundles",
    "cross_sell_rules",
    "upsell_rules",
    "product_catalog",
]


def clear_seeded_tables(conn: sqlite3.Connection) -> None:
    for table in SEEDED_TABLES:
        try:
            conn.execute(f"DELETE FROM {table}")
            print(f"  cleared: {table}")
        except sqlite3.OperationalError:
            pass
    conn.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    default_xlsx = Path.home() / "Desktop" / "Sales_Operating_System.xlsx"
    default_db = Path(__file__).parent.parent / "sales_os.db"

    parser = argparse.ArgumentParser(description="Seed Sales OS database from xlsx")
    parser.add_argument("--xlsx", default=str(default_xlsx), help="Path to xlsx workbook")
    parser.add_argument("--db", default=str(default_db), help="Path to SQLite database")
    parser.add_argument(
        "--clear", action="store_true", help="Clear seeded tables before inserting"
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    db_path = Path(args.db)

    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")
    if not db_path.exists():
        sys.exit(
            f"Database not found: {db_path}\n"
            "Run the application at least once to create the schema:\n"
            "  uvicorn app.main:app"
        )

    print(f"xlsx:  {xlsx_path}")
    print(f"db:    {db_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")

    if args.clear:
        print("\nClearing seeded tables...")
        clear_seeded_tables(conn)

    # name_to_id is shared across all seeders for FK resolution
    name_to_id: Dict[str, str] = {}

    print("\nSeeding...")
    seed_catalog(conn, wb, name_to_id)
    seed_upsell_rules(conn, wb, name_to_id)
    seed_cross_sell_rules(conn, wb, name_to_id)
    seed_bundles_and_items(conn, wb, name_to_id)
    seed_need_states(conn, wb, name_to_id)

    conn.close()
    print("\nSeed complete.")


if __name__ == "__main__":
    main()
