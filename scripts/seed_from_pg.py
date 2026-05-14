#!/usr/bin/env python3
"""
seed_from_pg.py
───────────────
Seeds the Sales Operating System Cloud SQL (Postgres) database from
Sales_Operating_System.xlsx, using the schema defined in
alembic/versions/001_initial_schema.py.

Key differences from seed_from_xlsx.py (SQLite edition):
  - Uses asyncpg (not sqlite3) via DATABASE_URL env var
  - UUID primary keys (Python-generated uuid.UUID objects)
  - JSONB columns receive Python dicts, not JSON strings
  - BOOLEAN columns receive True/False, not 1/0
  - TIMESTAMP WITH TIME ZONE columns receive timezone-aware datetime objects
  - Excel IDs (MC-001, etc.) stored in source_reference; fresh UUIDs for id
  - In-memory excel_id_to_uuid dict resolves all FK cross-references
  - Idempotent: skips rows where source_reference already exists
  - --clear flag: DELETE from tables in FK-safe order before inserting
  - --dry-run flag: parse xlsx and print what would be inserted, no DB calls

Usage:
    python scripts/seed_from_pg.py [--xlsx PATH] [--clear] [--dry-run]

Prerequisites:
    pip install asyncpg openpyxl
    export DATABASE_URL="postgresql://user:pass@host/dbname"
    # Cloud SQL Unix socket form (Cloud Run):
    # export DATABASE_URL="postgresql://user:pass@/dbname?host=/cloudsql/proj:region:inst"
"""

from __future__ import annotations

import argparse
import asyncio
import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

try:
    import asyncpg
except ImportError:
    sys.exit("asyncpg is required: pip install asyncpg")


# ─────────────────────────────────────────────────────────────────────────────
# DATABASE_URL normalization  (mirrors app/database.py)
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_dsn() -> str:
    raw = os.environ.get("DATABASE_URL", "")
    if not raw:
        sys.exit(
            "DATABASE_URL environment variable is not set.\n"
            "Example: export DATABASE_URL=postgresql://user:pass@host/dbname"
        )
    # asyncpg uses the bare postgresql:// scheme; strip SQLAlchemy driver tags
    dsn = raw
    for prefix in ("postgresql+asyncpg://", "postgres+asyncpg://"):
        if dsn.startswith(prefix):
            dsn = "postgresql://" + dsn[len(prefix):]
            break
    if dsn.startswith("postgres://"):
        dsn = "postgresql://" + dsn[len("postgres://"):]
    return dsn


# ─────────────────────────────────────────────────────────────────────────────
# Utilities
# ─────────────────────────────────────────────────────────────────────────────

def _now() -> datetime:
    return datetime.now(timezone.utc)


def _uid() -> uuid.UUID:
    return uuid.uuid4()


def _cell(val: Any) -> Optional[str]:
    if val is None:
        return None
    v = str(val).strip()
    return v if v else None


def _int_cell(val: Any, default: int = 1) -> int:
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def _float_cell(val: Any, default: float = 0.0) -> float:
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _sheet_rows(wb: openpyxl.Workbook, sheet_name: str) -> Tuple[List[str], List[Dict]]:
    """Return (headers, list_of_dicts) for a sheet, skipping empty rows.
    Identical logic to seed_from_xlsx.py."""
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    headers: List[str] = []
    rows: List[Dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not any(c for c in row):
            continue
        if not headers:
            headers = [
                str(c).strip() if c is not None else f"col_{j}"
                for j, c in enumerate(row)
            ]
            continue
        row_dict = {
            headers[j]: row[j]
            for j in range(min(len(headers), len(row)))
        }
        rows.append(row_dict)
    return headers, rows


def _name_to_slug(name: str) -> str:
    """Convert product name to stable lowercase slug for FK resolution."""
    return (
        name.lower().strip()
        .replace(" ", "-")
        .replace("/", "-")
        .replace("&", "and")
    )


# ─────────────────────────────────────────────────────────────────────────────
# FK-safe delete order (leaf tables first)
# ─────────────────────────────────────────────────────────────────────────────

SEEDED_TABLES_CLEAR_ORDER = [
    "need_state_products",
    "need_states",
    "bundle_items",
    "bundles",
    "cross_sell_rules",
    "upsell_rules",
    "product_catalog",
]

# Tables seeded from xlsx (we only seed these; clients/opportunities/etc. are
# populated by the application at runtime)
SEEDED_TABLES = list(SEEDED_TABLES_CLEAR_ORDER)


# ─────────────────────────────────────────────────────────────────────────────
# Seeders
# ─────────────────────────────────────────────────────────────────────────────

async def seed_catalog(
    conn: asyncpg.Connection,
    wb: openpyxl.Workbook,
    excel_id_to_uuid: Dict[str, uuid.UUID],
    dry_run: bool,
) -> None:
    """Seed product_catalog from Master_Catalog sheet."""
    _, rows = _sheet_rows(wb, "Master_Catalog")
    inserted = 0
    skipped = 0

    for row in rows:
        xlsx_id = _cell(row.get("ID"))
        name = _cell(row.get("Name"))
        if not name:
            continue

        new_id = _uid()
        slug = _name_to_slug(name)

        if dry_run:
            excel_id_to_uuid[slug] = new_id
            if xlsx_id:
                excel_id_to_uuid[xlsx_id.lower()] = new_id
            print(f"    [dry-run] product_catalog: {xlsx_id or name!r}")
            inserted += 1
            continue

        # Check idempotency by source_reference
        src_ref = xlsx_id or name
        existing = await conn.fetchrow(
            "SELECT id FROM product_catalog WHERE source_reference = $1",
            src_ref,
        )
        if existing:
            existing_id = existing["id"]
            excel_id_to_uuid[slug] = existing_id
            if xlsx_id:
                excel_id_to_uuid[xlsx_id.lower()] = existing_id
            skipped += 1
            continue

        excel_id_to_uuid[slug] = new_id
        if xlsx_id:
            excel_id_to_uuid[xlsx_id.lower()] = new_id

        await conn.execute(
            """
            INSERT INTO product_catalog (
                id, name, type, category, subcategory, description,
                primary_goal, core_value, interaction_value,
                marketing_influence, score_multiplier, funnel_stage,
                primary_channel, automation_potential, source_reference,
                is_active, created_at, updated_at
            ) VALUES (
                $1, $2, $3, $4, $5, $6,
                $7, $8, $9,
                $10, $11, $12,
                $13, $14, $15,
                $16, $17, $18
            )
            """,
            new_id,
            name,
            _cell(row.get("Type")),
            _cell(row.get("Category")),
            _cell(row.get("Subcategory")),
            _cell(row.get("Description")),
            _cell(row.get("Primary_Goal")),
            _cell(row.get("Core_Value")),
            _int_cell(row.get("Interaction_Value (1-5)"), 3),
            _int_cell(row.get("Marketing_Influence (1-5)"), 3),
            _float_cell(row.get("Score_Multiplier"), 1.0),
            _cell(row.get("Funnel_Stage")),
            _cell(row.get("Primary_Channel")),
            _cell(row.get("Automation_Potential")),
            src_ref,
            True,          # is_active — BOOLEAN
            _now(),        # created_at — TIMESTAMPTZ
            _now(),        # updated_at — TIMESTAMPTZ
        )
        inserted += 1

    print(f"  product_catalog: {inserted} inserted, {skipped} skipped", flush=True)


async def seed_upsell_rules(
    conn: asyncpg.Connection,
    wb: openpyxl.Workbook,
    excel_id_to_uuid: Dict[str, uuid.UUID],
    dry_run: bool,
) -> None:
    """Seed upsell_rules from Upsell_Matrix sheet."""
    _, rows = _sheet_rows(wb, "Upsell_Matrix")
    inserted = 0

    for row in rows:
        primary_name = _cell(row.get("Primary_Product"))
        rec_name = _cell(row.get("Recommended_Upsell"))
        dep_name = _cell(row.get("Dependency"))

        if not primary_name or not rec_name:
            continue

        primary_id = excel_id_to_uuid.get(_name_to_slug(primary_name))
        rec_id = excel_id_to_uuid.get(_name_to_slug(rec_name))
        dep_id = excel_id_to_uuid.get(_name_to_slug(dep_name)) if dep_name else None

        new_id = _uid()

        if dry_run:
            print(
                f"    [dry-run] upsell_rules: "
                f"{primary_name!r} → {rec_name!r}"
            )
            inserted += 1
            continue

        await conn.execute(
            """
            INSERT INTO upsell_rules (
                id, primary_product_id, trigger_event,
                client_need_state_id, recommended_product_id,
                upsell_type, expected_impact, dependency_product_id
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            """,
            new_id,
            primary_id,
            _cell(row.get("Trigger_Event")),
            None,   # client_need_state_id — resolved after need_states seeded
            rec_id,
            _cell(row.get("Upsell_Type")),
            _cell(row.get("Expected_Impact")),
            dep_id,
        )
        inserted += 1

    print(f"  upsell_rules: {inserted} inserted", flush=True)


async def seed_cross_sell_rules(
    conn: asyncpg.Connection,
    wb: openpyxl.Workbook,
    excel_id_to_uuid: Dict[str, uuid.UUID],
    dry_run: bool,
) -> None:
    """Seed cross_sell_rules from Cross_Sell_Matrix sheet."""
    _, rows = _sheet_rows(wb, "Cross_Sell_Matrix")
    inserted = 0

    for row in rows:
        prod_name = _cell(row.get("Product"))
        paired_name = _cell(row.get("Pairs_With"))
        if not prod_name or not paired_name:
            continue

        prod_id = excel_id_to_uuid.get(_name_to_slug(prod_name))
        paired_id = excel_id_to_uuid.get(_name_to_slug(paired_name))

        if dry_run:
            print(
                f"    [dry-run] cross_sell_rules: "
                f"{prod_name!r} ↔ {paired_name!r}"
            )
            inserted += 1
            continue

        await conn.execute(
            """
            INSERT INTO cross_sell_rules (
                id, product_id, paired_product_id, reason, bundle_strength
            ) VALUES ($1, $2, $3, $4, $5)
            """,
            _uid(),
            prod_id,
            paired_id,
            _cell(row.get("Reason")),
            _int_cell(row.get("Bundle_Strength (1-5)"), 3),
        )
        inserted += 1

    print(f"  cross_sell_rules: {inserted} inserted", flush=True)


async def seed_bundles_and_items(
    conn: asyncpg.Connection,
    wb: openpyxl.Workbook,
    excel_id_to_uuid: Dict[str, uuid.UUID],
    dry_run: bool,
) -> None:
    """Seed bundles and bundle_items from Bundles sheet."""
    _, rows = _sheet_rows(wb, "Bundles")
    bundle_name_to_uuid: Dict[str, uuid.UUID] = {}
    bundle_sequence: Dict[str, int] = {}
    bundles_inserted = 0
    items_inserted = 0

    for row in rows:
        bundle_name = _cell(row.get("Bundle_Name"))
        component_name = _cell(row.get("Component"))
        if not bundle_name or not component_name:
            continue

        # Create bundle if new
        if bundle_name not in bundle_name_to_uuid:
            if dry_run:
                bundle_id = _uid()
                bundle_name_to_uuid[bundle_name] = bundle_id
                bundle_sequence[bundle_name] = 0
                print(f"    [dry-run] bundles: {bundle_name!r}")
                bundles_inserted += 1
            else:
                existing = await conn.fetchrow(
                    "SELECT id FROM bundles WHERE name = $1", bundle_name
                )
                if existing:
                    bundle_name_to_uuid[bundle_name] = existing["id"]
                else:
                    bundle_id = _uid()
                    target_need = _cell(row.get("Target_Need"))
                    await conn.execute(
                        """
                        INSERT INTO bundles (
                            id, name, value_proposition, created_at, updated_at
                        ) VALUES ($1, $2, $3, $4, $5)
                        """,
                        bundle_id,
                        bundle_name,
                        target_need,
                        _now(),
                        _now(),
                    )
                    bundle_name_to_uuid[bundle_name] = bundle_id
                    bundle_sequence[bundle_name] = 0
                    bundles_inserted += 1

        bundle_id = bundle_name_to_uuid[bundle_name]
        product_id = excel_id_to_uuid.get(_name_to_slug(component_name))
        if not product_id:
            continue

        seq = bundle_sequence.get(bundle_name, 0)
        bundle_sequence[bundle_name] = seq + 1

        if dry_run:
            print(
                f"    [dry-run] bundle_items: "
                f"{bundle_name!r} / {component_name!r}"
            )
            items_inserted += 1
            continue

        # Skip duplicate bundle_items
        existing_item = await conn.fetchrow(
            "SELECT id FROM bundle_items WHERE bundle_id = $1 AND product_id = $2",
            bundle_id, product_id,
        )
        if existing_item:
            continue

        await conn.execute(
            """
            INSERT INTO bundle_items (
                id, bundle_id, product_id, sequence_order, required
            ) VALUES ($1, $2, $3, $4, $5)
            """,
            _uid(),
            bundle_id,
            product_id,
            seq,
            True,   # required — BOOLEAN
        )
        items_inserted += 1

    print(f"  bundles: {bundles_inserted} inserted", flush=True)
    print(f"  bundle_items: {items_inserted} inserted", flush=True)


async def seed_need_states(
    conn: asyncpg.Connection,
    wb: openpyxl.Workbook,
    excel_id_to_uuid: Dict[str, uuid.UUID],
    dry_run: bool,
) -> None:
    """Seed need_states and need_state_products from Client_Needs_Mapping sheet."""
    _, rows = _sheet_rows(wb, "Client_Needs_Mapping")
    ns_inserted = 0
    nsp_inserted = 0

    for row in rows:
        problem = _cell(row.get("Client_Problem"))
        signal = _cell(row.get("Detected_Signal"))
        products_raw = _cell(row.get("Recommended_Products"))
        priority = _int_cell(row.get("Priority_Order"), 2)

        if not problem:
            continue

        if dry_run:
            print(f"    [dry-run] need_states: {problem!r}")
            ns_id = _uid()
            ns_inserted += 1
        else:
            existing = await conn.fetchrow(
                "SELECT id FROM need_states WHERE problem_name = $1", problem
            )
            if existing:
                ns_id = existing["id"]
            else:
                ns_id = _uid()
                await conn.execute(
                    """
                    INSERT INTO need_states (id, problem_name, detected_signal)
                    VALUES ($1, $2, $3)
                    """,
                    ns_id,
                    problem,
                    signal,
                )
                ns_inserted += 1

        if products_raw:
            for prod_name in [p.strip() for p in products_raw.split(";")]:
                if not prod_name:
                    continue
                prod_id = excel_id_to_uuid.get(_name_to_slug(prod_name))
                if not prod_id:
                    continue

                if dry_run:
                    print(
                        f"    [dry-run] need_state_products: "
                        f"{problem!r} → {prod_name!r}"
                    )
                    nsp_inserted += 1
                    continue

                existing_link = await conn.fetchrow(
                    "SELECT id FROM need_state_products "
                    "WHERE need_state_id = $1 AND product_id = $2",
                    ns_id, prod_id,
                )
                if existing_link:
                    continue

                await conn.execute(
                    """
                    INSERT INTO need_state_products (
                        id, need_state_id, product_id, priority_order
                    ) VALUES ($1, $2, $3, $4)
                    """,
                    _uid(),
                    ns_id,
                    prod_id,
                    priority,
                )
                nsp_inserted += 1

    print(f"  need_states: {ns_inserted} inserted", flush=True)
    print(f"  need_state_products: {nsp_inserted} inserted", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Clear helpers
# ─────────────────────────────────────────────────────────────────────────────

async def clear_seeded_tables(conn: asyncpg.Connection) -> None:
    """DELETE from seeded tables in FK-safe order (leaf tables first)."""
    for table in SEEDED_TABLES_CLEAR_ORDER:
        try:
            result = await conn.execute(f"DELETE FROM {table}")
            print(f"  cleared: {table} ({result})", flush=True)
        except asyncpg.UndefinedTableError:
            print(f"  skip (table not found): {table}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

async def _run(xlsx_path: Path, clear: bool, dry_run: bool) -> None:
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")

    print(f"xlsx:  {xlsx_path}", flush=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # excel_id_to_uuid is shared across all seeders for FK resolution.
    # Keys: lower-cased Excel ID (e.g. "mc-001") OR name slug.
    # Values: uuid.UUID objects (asyncpg's native UUID type).
    excel_id_to_uuid: Dict[str, uuid.UUID] = {}

    if dry_run:
        print("\n[dry-run] Parsing xlsx — no DB calls will be made.\n")
        conn = None
    else:
        dsn = _resolve_dsn()
        print(f"dsn:   {dsn[:40]}...", flush=True)
        conn = await asyncpg.connect(dsn)
        # Register UUID codec so asyncpg returns uuid.UUID objects natively
        await conn.set_type_codec(
            "uuid",
            encoder=str,
            decoder=uuid.UUID,
            schema="pg_catalog",
            format="text",
        )

    try:
        if clear and not dry_run:
            print("\nClearing seeded tables...")
            await clear_seeded_tables(conn)

        print("\nSeeding...")
        if dry_run:
            # Dry-run path: no real connection needed
            await seed_catalog(None, wb, excel_id_to_uuid, dry_run=True)
            await seed_upsell_rules(None, wb, excel_id_to_uuid, dry_run=True)
            await seed_cross_sell_rules(None, wb, excel_id_to_uuid, dry_run=True)
            await seed_bundles_and_items(None, wb, excel_id_to_uuid, dry_run=True)
            await seed_need_states(None, wb, excel_id_to_uuid, dry_run=True)
        else:
            async with conn.transaction():
                await seed_catalog(conn, wb, excel_id_to_uuid, dry_run=False)
                await seed_upsell_rules(conn, wb, excel_id_to_uuid, dry_run=False)
                await seed_cross_sell_rules(conn, wb, excel_id_to_uuid, dry_run=False)
                await seed_bundles_and_items(conn, wb, excel_id_to_uuid, dry_run=False)
                await seed_need_states(conn, wb, excel_id_to_uuid, dry_run=False)
    finally:
        if conn is not None:
            await conn.close()

    print("\nSeed complete.", flush=True)


def main() -> None:
    default_xlsx = Path.home() / "Desktop" / "Sales_Operating_System.xlsx"

    parser = argparse.ArgumentParser(
        description="Seed Sales OS Postgres database from xlsx"
    )
    parser.add_argument(
        "--xlsx",
        default=str(default_xlsx),
        help="Path to Sales_Operating_System.xlsx "
             "(default: ~/Desktop/Sales_Operating_System.xlsx)",
    )
    parser.add_argument(
        "--clear",
        action="store_true",
        help="DELETE from seeded tables in FK-safe order before inserting",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse xlsx and print what would be inserted; no DB calls",
    )
    args = parser.parse_args()

    asyncio.run(_run(
        xlsx_path=Path(args.xlsx),
        clear=args.clear,
        dry_run=args.dry_run,
    ))


if __name__ == "__main__":
    main()
